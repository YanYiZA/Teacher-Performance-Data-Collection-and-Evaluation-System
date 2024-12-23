# -*- coding: utf-8 -*-

import sys
import os
from openpyxl import Workbook
from docx import Document

# 导出到Excel
def export_to_excel(output_file):
    wb = Workbook()
    ws = wb.active

    # 表头
    columns = ["姓名", "教学工作", "教研工作", "科研工作", "其他工作", "总绩效"]
    for col_num, header in enumerate(columns, start=1):
        ws.cell(row=1, column=col_num, value=header)

    # 读取数据文件
    with open('data.txt', 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # 写入数据
    for row_num, line in enumerate(lines, start=2):
        cells = line.strip().split('\t')
        for col_num, cell in enumerate(cells, start=1):
            ws.cell(row=row_num, column=col_num, value=cell)

    wb.save(output_file)
    print(f"Excel 文件已保存: {output_file}")

# 导出到Word
def export_to_word(output_file):
    doc = Document()
    columns = ["姓名", "教学工作", "教研工作", "科研工作", "其他工作", "总绩效"]

    # 读取数据文件
    with open('data.txt', 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # 创建表格并填充表头
    table = doc.add_table(rows=1, cols=len(columns))
    hdr_cells = table.rows[0].cells
    for i, header in enumerate(columns):
        hdr_cells[i].text = header

    # 写入数据
    for line in lines:
        row_cells = table.add_row().cells
        cells = line.strip().split('\t')
        for i, cell in enumerate(cells):
            row_cells[i].text = cell

    doc.save(output_file)
    print(f"Word 文件已保存: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("用法: python export_data.py <格式> <输出文件>")
        sys.exit(1)

    file_format = sys.argv[1].lower()
    output_file = sys.argv[2]

    if not os.path.exists('data.txt'):
        print("错误: 数据文件 data.txt 不存在。")
        sys.exit(1)

    if file_format == 'excel':
        export_to_excel(output_file)
    elif file_format == 'word':
        export_to_word(output_file)
    else:
        print("错误: 不支持的格式，请指定 'excel' 或 'word'。")
        sys.exit(1)
